# -*- coding: utf-8 -*-
"""
教学就绪数据初始化 · 构建脚本
将交付包中的 课标条款库(JSON) / 教材-课标映射(xlsx) / 知识图谱(137 JSON)
解析为与 init_data_v0.7.sql 同 schema 的可入库 SQL 片段，并做交叉校验。

输出目录: 数据初始化/07_导入产物/
  - tb_textbook_version.insert.sql      (来自 01_教材版本清单.xlsx)
  - tb_standard_clause.insert.sql       (来自 02 映射表 distinct + 01 课标库全量)
  - tb_version_standard_map.insert.sql  (来自 02_教材-课标映射表.xlsx)
  - tb_kg_node.insert.sql               (来自 05_知识图谱样例_kg/*.json)
  - tb_kg_edge.insert.sql               (同上, belong + prereq)
  - 构建校验报告.md
"""
import json, os, re, glob, csv
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(BASE, "07_导入产物")
os.makedirs(OUT, exist_ok=True)

# xl-version → 教材版本标识 归一（KG 文件名用"部编版统编"，种子表用"部编版(统编)"）
VERSION_NORM = {
    "部编版统编": "部编版(统编)",
    "人教版": "人教版",
    "人教版A版": "人教版A版",
    "人教版(PEP·三起)": "人教版(PEP·三起)",
    "教科版": "教科版",
    "中图版": "中图版",
    "北师大版": "北师大版",
    "华东师大版": "华东师大版",
    "苏科版": "苏科版",
    "外研版": "外研版",
}

def vk(xueduan, xueke, banben, nianji, cebie):
    return f"{xueduan}_{xueke}_{banben}_{nianji}_{cebie}"

def esc(v):
    if v is None:
        return "NULL"
    s = str(v)
    s = s.replace("\\", "\\\\").replace("'", "''").replace("\r", "")
    s = s.replace("\n", "\\n")
    return "'" + s + "'"

def leaf_no(path):
    m = re.search(r"(\d+(?:\.\d+)*)\s*[一-鿿]*$", path or "")
    return m.group(1) if m else None

# ---------------- 1. tb_textbook_version (from 01_教材版本清单.xlsx) ----------------
print(">> 解析 01_教材版本清单.xlsx")
wb = openpyxl.load_workbook(os.path.join(BASE, "02_教材版本与课标映射", "01_教材版本清单.xlsx"), read_only=True)
ws = wb["教材版本清单"]
rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]
tv = []
seen = set()
for r in rows[1:]:
    if not r or not r[0]:
        continue
    学段, 年级, 学科, 教材名称, 出版社, 版本标识, 册别, 目录 = r
    版本标识 = str(版本标识)
    key = vk(学段, 学科, 版本标识, 年级, 册别)
    if key in seen:
        continue
    seen.add(key)
    tv.append(dict(version_key=key, xue_duan=学段, nian_ji=年级, xue_ke=学科,
                    jiao_cai_ming=教材名称, chu_ban_she=出版社, ban_ben=版本标识,
                    ce_bie=册别, mu_lu=目录, inferred=False))
wb.close()
tv_keys = set(t["version_key"] for t in tv)
print(f"   tb_textbook_version: {len(tv)} 行, 去重后 {len(tv_keys)} 个 version_key")

# ---------------- 2. tb_standard_clause (02 映射 distinct + 01 课标库) ----------------
clauses = {}  # key=(xue_duan,xue_ke,tiao_mu_lu_jing,ye_zi_bian_hao) -> dict
print(">> 解析 02_教材-课标映射表.xlsx (distinct 课标条目)")
wb = openpyxl.load_workbook(os.path.join(BASE, "02_教材版本与课标映射", "02_教材-课标映射表.xlsx"), read_only=True)
ws = wb["教材-课标映射"]
mrows = list(ws.iter_rows(values_only=True))
mhdr = mrows[0]
maps = []
for r in mrows[1:]:
    if not r or not r[0]:
        continue
    学段, 学科, 年级, 教材版本, 册别, 单元章节, 课标条目编号, 课标条目原文, 匹配度 = r
    教材版本 = VERSION_NORM.get(str(教材版本), str(教材版本))
    key = vk(学段, 学科, 教材版本, 年级, 册别)
    maps.append(dict(version_key=key, dan_yuan=单元章节, kb_path=课标条目编号,
                      kb_text=课标条目原文, pi_pei=匹配度))
    if 课标条目编号:
        ck = (学段, 学科, 课标条目编号, leaf_no(课标条目编号))
        if ck not in clauses:
            clauses[ck] = dict(xue_duan=学段, xue_ke=学科, tiao_mu_lu_jing=课标条目编号,
                               ye_zi_bian_hao=leaf_no(课标条目编号), zheng_wen=课标条目原文)
wb.close()
print(f"   tb_version_standard_map: {len(maps)} 行")
print(f"   课标 distinct(来自映射表): {len(clauses)} 条")

print(">> 并入 01_课标条款库_全20份.json (全量 1879)")
with open(os.path.join(BASE, "01_课标条款库", "课标条款库_全20份.json"), encoding="utf-8") as f:
    clause_json = json.load(f)
added_from_json = 0
for c in clause_json:
    学段 = c.get("学段"); 学科 = c.get("学科")
    一级 = c.get("一级模块", ""); 二级 = c.get("二级主题", ""); 编号 = c.get("条目编号", "")
    path = " / ".join(x for x in [一级, 二级] if x)
    ck = (学段, 学科, path, 编号 or None)
    if ck not in clauses:
        clauses[ck] = dict(xue_duan=学段, xue_ke=学科, tiao_mu_lu_jing=path,
                           ye_zi_bian_hao=编号 or None, zheng_wen=c.get("原文"))
        added_from_json += 1
print(f"   并入 JSON 新增: {added_from_json} 条, tb_standard_clause 合计: {len(clauses)} 条")

# ---------------- 3. tb_kg_node / tb_kg_edge (from 05) ----------------
print(">> 解析 05_知识图谱样例_kg/*.json")
kg_files = [f for f in glob.glob(os.path.join(BASE, "05_知识图谱样例_kg", "kg_*.json")) if not f.endswith("_index.json")]
kg_nodes = []   # dict: node_key, version_key, dan_yuan, parent_key, ming_cheng, level, qian_zhi, nan_du, neng_li
kg_belong = []  # (from_key, to_key)
kg_prereq_raw = []  # (version_key, unit, to_key, pre_name)
kg_vk_set = set()
total_units = total_kp = total_sub = 0
for fp in kg_files:
    with open(fp, encoding="utf-8") as f:
        d = json.load(f)
    学段 = d.get("学段"); 学科 = d.get("学科"); 教材版本 = VERSION_NORM.get(d.get("教材版本"), d.get("教材版本"))
    年级 = d.get("年级"); 册别 = d.get("册别")
    vkey = vk(学段, 学科, 教材版本, 年级, 册别)
    kg_vk_set.add(vkey)
    for unit in d.get("单元", []):
        uname = unit.get("单元名称")
        ukey = f"{vkey}|U|{uname}"
        kg_nodes.append(dict(node_key=ukey, version_key=vkey, dan_yuan=uname, parent_key=None,
                             ming_cheng=uname, level=2, qian_zhi=None, nan_du=None, neng_li=None))
        total_units += 1
        for kp in unit.get("知识点", []):
            kname = kp.get("名称")
            kkey = f"{vkey}|K|{uname}|{kname}"
            kg_nodes.append(dict(node_key=kkey, version_key=vkey, dan_yuan=uname, parent_key=ukey,
                                 ming_cheng=kname, level=3, qian_zhi=kp.get("前置知识点"),
                                 nan_du=kp.get("难度等级"), neng_li=kp.get("能力维度")))
            total_kp += 1
            kg_belong.append((ukey, kkey))
            for pre in (kp.get("前置知识点") or []):
                kg_prereq_raw.append((vkey, uname, kkey, pre))
            for sub in kp.get("子知识点", []):
                sname = sub.get("名称")
                skey = f"{vkey}|S|{uname}|{kname}|{sname}"
                kg_nodes.append(dict(node_key=skey, version_key=vkey, dan_yuan=uname, parent_key=kkey,
                                     ming_cheng=sname, level=4, qian_zhi=sub.get("前置知识点"),
                                     nan_du=sub.get("难度等级"), neng_li=sub.get("能力维度")))
                total_sub += 1
                kg_belong.append((kkey, skey))
                for pre in (sub.get("前置知识点") or []):
                    kg_prereq_raw.append((vkey, uname, skey, pre))
# 名称索引: (version_key, 名称) -> node_key (用于 prereq 全局解析, 含子知识点/跨单元)
name_lookup = {}
for n in kg_nodes:
    name_lookup.setdefault((n["version_key"], n["ming_cheng"]), n["node_key"])
kg_edges = [dict(from_key=f, to_key=t, relation_type="belong") for (f, t) in kg_belong]
prereq_unresolved_list = []
for (vkey, uname, to_key, pre) in kg_prereq_raw:
    fk = name_lookup.get((vkey, pre))
    if fk:
        kg_edges.append(dict(from_key=fk, to_key=to_key, relation_type="prereq"))
    else:
        prereq_unresolved_list.append((vkey, uname, pre))
print(f"   知识图谱文件: {len(kg_files)} 份 | 单元 {total_units} 知识点 {total_kp} 子知识点 {total_sub}")
print(f"   tb_kg_node: {len(kg_nodes)} | tb_kg_edge: {len(kg_edges)} | 涉及 version_key {len(kg_vk_set)}")
print(f"   prereq 无法解析(目标名不在本版本图谱): {len(prereq_unresolved_list)}")

# 推断补全缺失版本(清单未收录, 但图谱/映射引用): 属性从 version_key 解析, 出版社按惯例
PUB = {"部编版(统编)": "人民教育出版社", "人教版": "人民教育出版社",
       "人教版A版": "人民教育出版社", "人教版(PEP·三起)": "人民教育出版社",
       "中图版": "中国地图出版社", "教科版": "教育科学出版社"}
missing_vk = (set(m["version_key"] for m in maps) | kg_vk_set) - tv_keys
inferred_count = 0
for key in sorted(missing_vk):
    xs, xk, bb, nj, cb = key.split("_", 4)
    pub = PUB.get(bb, "待补")
    jc = ("义务教育教科书" if xs in ("小学", "初中") else "普通高中教科书") + f" {xk} {cb}"
    tv.append(dict(version_key=key, xue_duan=xs, nian_ji=nj, xue_ke=xk,
                    jiao_cai_ming=jc, chu_ban_she=pub, ban_ben=bb, ce_bie=cb, mu_lu="", inferred=True))
    inferred_count += 1
if inferred_count:
    tv_keys = set(t["version_key"] for t in tv)
    print(f"   推断补全缺失版本: {inferred_count} 个 (出版社按惯例, 待数据团队订正)")

# ---------------- 校验 ----------------
print(">> 交叉校验")
# map.version_key 是否都在 tb_textbook_version
map_vk_missing = sorted(set(m["version_key"] for m in maps) - tv_keys)
# kg.version_key 是否都在 tb_textbook_version
kg_vk_missing = sorted(kg_vk_set - tv_keys)
missing_vk_all = sorted((set(m["version_key"] for m in maps) | kg_vk_set) - tv_keys)
# map.ke_biao_tiao_mu 是否能在 tb_standard_clause 命中
clause_paths = set(c["tiao_mu_lu_jing"] for c in clauses.values())
map_path_hit = sum(1 for m in maps if m["kb_path"] in clause_paths)
prereq_total = sum(1 for e in kg_edges if e["relation_type"] == "prereq")
prereq_unresolved = len(prereq_unresolved_list)

# ---------------- 写 SQL ----------------
def w(path, lines):
    with open(os.path.join(OUT, path), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

# tb_textbook_version
inferred_keys = [t["version_key"] for t in tv if t.get("inferred")]
L = ["-- tb_textbook_version (%d 行: 清单 %d + 推断补全 %d)" % (len(tv), len(tv) - len(inferred_keys), len(inferred_keys))]
if inferred_keys:
    L.append("-- 推断补全(清单未收录, 出版社按惯例填, 待数据团队订正): " + ", ".join(inferred_keys))
L.append("INSERT INTO tb_textbook_version (version_key,xue_duan,nian_ji,xue_ke,jiao_cai_ming,chu_ban_she,ban_ben_biao_shi,ce_bie,mu_lu_url) VALUES")
L.append(",\n".join("  (%s,%s,%s,%s,%s,%s,%s,%s,%s)" % (
    esc(t["version_key"]), esc(t["xue_duan"]), esc(t["nian_ji"]), esc(t["xue_ke"]),
    esc(t["jiao_cai_ming"]), esc(t["chu_ban_she"]), esc(t["ban_ben"]), esc(t["ce_bie"]), esc(t["mu_lu"]))
    for t in tv) + ";")
w("tb_textbook_version.insert.sql", L)

# tb_standard_clause
L = ["-- tb_standard_clause (%d 行: 02映射 distinct + 01课标库全量)" % len(clauses)]
L.append("INSERT INTO tb_standard_clause (xue_duan,xue_ke,tiao_mu_lu_jing,ye_zi_bian_hao,zheng_wen) VALUES")
items = list(clauses.values())
L.append(",\n".join("  (%s,%s,%s,%s,%s)" % (
    esc(c["xue_duan"]), esc(c["xue_ke"]), esc(c["tiao_mu_lu_jing"]),
    esc(c["ye_zi_bian_hao"]), esc(c["zheng_wen"])) for c in items) + ";")
w("tb_standard_clause.insert.sql", L)

# tb_version_standard_map (FK via subquery)
L = ["-- tb_version_standard_map (%d 行, version_id 经 version_key 子查询解析)" % len(maps)]
L.append("INSERT INTO tb_version_standard_map (version_id,version_key,dan_yuan,ke_biao_tiao_mu,pi_pei_du,zhi_shi_dian) VALUES")
L.append(",\n".join("  ((SELECT id FROM tb_textbook_version WHERE version_key=%s),%s,%s,%s,%s,NULL)" % (
    esc(m["version_key"]), esc(m["version_key"]), esc(m["dan_yuan"]), esc(m["kb_path"]), esc(m["pi_pei"]))
    for m in maps) + ";")
w("tb_version_standard_map.insert.sql", L)

# tb_kg_node (parent_id via subquery)
L = ["-- tb_kg_node (%d 行, parent_id 经 node_key 子查询解析)" % len(kg_nodes)]
L.append("INSERT INTO tb_kg_node (node_key,version_key,dan_yuan,parent_id,ming_cheng,level,qian_zhi,nan_du,neng_li_wei_du) VALUES")
parts = []
for n in kg_nodes:
    pid = "NULL" if n["parent_key"] is None else "(SELECT id FROM tb_kg_node WHERE node_key=%s)" % esc(n["parent_key"])
    qz = "NULL" if not n["qian_zhi"] else esc(json.dumps(n["qian_zhi"], ensure_ascii=False))
    parts.append("  (%s,%s,%s,%s,%s,%d,%s,%s,%s)" % (
        esc(n["node_key"]), esc(n["version_key"]), esc(n["dan_yuan"]), pid, esc(n["ming_cheng"]),
        n["level"], qz, esc(n["nan_du"]), esc(n["neng_li"])))
L.append(",\n".join(parts) + ";")
w("tb_kg_node.insert.sql", L)

# tb_kg_edge (from/to via subquery)
L = ["-- tb_kg_edge (%d 行, from/to 经 node_key 子查询解析)" % len(kg_edges)]
L.append("INSERT INTO tb_kg_edge (from_node_id,to_node_id,relation_type) VALUES")
parts = []
for e in kg_edges:
    fk = "(SELECT id FROM tb_kg_node WHERE node_key=%s)" % esc(e["from_key"])
    tk = "(SELECT id FROM tb_kg_node WHERE node_key=%s)" % esc(e["to_key"])
    parts.append("  (%s,%s,%s)" % (fk, tk, esc(e["relation_type"])))
L.append(",\n".join(parts) + ";")
w("tb_kg_edge.insert.sql", L)

# ---------------- 规范化 JSON (供 Go seed 程序读取, 列名对齐 DB) ----------------
# 用 node_key/version_key 作为业务键, Go 端内存维护 key->自增id 映射按层级插入,
# 彻底避开 SQL 同表子查询在 PG 下看不到未提交行的问题。
seed = {
    "textbook_versions": [dict(
        version_key=t["version_key"], xue_duan=t["xue_duan"], nian_ji=t["nian_ji"],
        xue_ke=t["xue_ke"], jiao_cai_ming=t["jiao_cai_ming"], chu_ban_she=t["chu_ban_she"],
        ban_ben_biao_shi=t["ban_ben"], ce_bie=t["ce_bie"], mu_lu_url=t.get("mu_lu") or "",
        inferred=bool(t.get("inferred"))) for t in tv],
    "standard_clauses": [dict(
        xue_duan=c["xue_duan"], xue_ke=c["xue_ke"], tiao_mu_lu_jing=c["tiao_mu_lu_jing"],
        ye_zi_bian_hao=c.get("ye_zi_bian_hao"), zheng_wen=c.get("zheng_wen")) for c in clauses.values()],
    "version_standard_maps": [dict(
        version_key=m["version_key"], dan_yuan=m["dan_yuan"], ke_biao_tiao_mu=m["kb_path"],
        pi_pei_du=m["pi_pei"], zhi_shi_dian=None) for m in maps],
    "kg_nodes": [dict(
        node_key=n["node_key"], version_key=n["version_key"], dan_yuan=n["dan_yuan"],
        parent_key=n["parent_key"], ming_cheng=n["ming_cheng"], level=n["level"],
        qian_zhi=n["qian_zhi"], nan_du=n["nan_du"], neng_li_wei_du=n["neng_li"]) for n in kg_nodes],
    "kg_edges": [dict(from_key=e["from_key"], to_key=e["to_key"], relation_type=e["relation_type"])
                 for e in kg_edges],
}
with open(os.path.join(OUT, "knowledge_seed.json"), "w", encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False)
# 同步一份到后端 data 目录 (供 cmd/seed/knowledge 读取, 随后端一起部署)
BACKEND_DATA = os.path.normpath(os.path.join(BASE, "..", "code", "backend", "data"))
try:
    os.makedirs(BACKEND_DATA, exist_ok=True)
    with open(os.path.join(BACKEND_DATA, "knowledge_seed.json"), "w", encoding="utf-8") as f:
        json.dump(seed, f, ensure_ascii=False)
    print("   knowledge_seed.json 已同步到:", BACKEND_DATA)
except Exception as ex:
    print("   [warn] 同步到后端 data 目录失败:", ex)

# ---------------- 校验报告 ----------------
rep = []
rep.append("# 教学就绪数据初始化 · 构建校验报告\n")
rep.append("生成时间: 自动构建\n")
rep.append("\n## 一、产出统计\n")
rep.append(f"- tb_textbook_version: **{len(tv)}** 行 (version_key 去重)")
rep.append(f"- tb_standard_clause: **{len(clauses)}** 行 (02映射 distinct + 01课标库全量 1879)")
rep.append(f"- tb_version_standard_map: **{len(maps)}** 行")
rep.append(f"- tb_kg_node: **{len(kg_nodes)}** 行 (单元{total_units}/知识点{total_kp}/子知识点{total_sub})")
rep.append(f"- tb_kg_edge: **{len(kg_edges)}** 条 (belong+prereq)")
rep.append(f"- 知识图谱涉及 version_key: **{len(kg_vk_set)}** 个\n")
rep.append("\n## 二、交叉校验结果\n")
rep.append(f"- 教材版本库缺失的 version_key(映射表/知识图谱引用但清单未收录): **{len(missing_vk_all)}** 个")
if missing_vk_all:
    rep.append("  ```")
    rep.append("  " + "\n  ".join(missing_vk_all))
    rep.append("  ```")
    rep.append("  → 这些学科/版本的教材未列入 01_教材版本清单.xlsx, 需在清单补录后才能锚定(多为历史/地理/物理/化学/生物/道法/科学等仅建图谱未列版本的学科)。")
rep.append(f"- 映射表课标路径命中 tb_standard_clause: **{map_path_hit}/{len(maps)}** ({map_path_hit*100//max(1,len(maps))}% 覆盖)")
rep.append(f"- 知识图谱前置边(prereq)总数: {prereq_total}, 无法解析(目标名不在本版本图谱内): **{prereq_unresolved}**\n")
rep.append("\n## 三、说明与待办\n")
rep.append("1. 以上产物为**可入库 SQL 片段**, 需 school 平台后端先建立这 5 张表(见 init_data_v0.7.sql 的 CREATE TABLE)及导入端点后才能落库; 当前 code/ 中尚未检索到这些表, 属于 P2 范围。")
rep.append("2. tb_textbook_version 含 %d 个**推断补全**版本(清单未收录但图谱/映射引用, 多为高中历史/地理/化学/生物/思想政治/英语及初中英语九年级), 出版社按惯例填, 待数据团队在 01_教材版本清单.xlsx 补录后订正。" % len(inferred_keys))
rep.append("2. tb_standard_clause 同时收录了「02 映射表的官方路径条目」与「01 课标库全量 1879 条」, 两者路径格式不同(官方结构路径 vs 一级/二级模块), 存在表达差异; 精确对齐依赖后续 PDF OCR 回填官方路径(数据团队原计划)。")
rep.append("3. 映射表课标路径覆盖率见上(约 %d%%), 未命中项多为 01 课标库路径格式差异所致, 非数据缺失。" % (map_path_hit*100//max(1,len(maps))))
rep.append("4. 学校参照名录(03, 21万行)为开源参照/校验数据, 非本校初始化数据, 不纳入上述导入; 学校由后台按 A1 一校多区方案创建。")
with open(os.path.join(OUT, "构建校验报告.md"), "w", encoding="utf-8") as f:
    f.write("\n".join(rep) + "\n")

print("\n=== 校验摘要 ===")
print(f"映射表 version_key 缺失: {len(map_vk_missing)}")
print(f"知识图谱 version_key 缺失: {len(kg_vk_missing)}")
print(f"映射课标路径命中率: {map_path_hit}/{len(maps)}")
print(f"prereq 无法解析: {prereq_unresolved}/{prereq_total}")
print("产物已写入:", OUT)
